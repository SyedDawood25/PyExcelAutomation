from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Syed Dawood" : {
        "Programming" : 100,
        "Mathematics" : 99,
        "ICT" : 100
    },
    "Moiz Habib" : {
        "Programming" : 90,
        "Mathematics" : 99,
        "ICT" : 95
    },
    "Saad Rasheed" : {
        "Programming" : 50,
        "Mathematics" : 60,
        "ICT" : 70
    }
}

wb = Workbook()
ws = wb.active
ws.title = 'Grades'

headings = ['Name'] + list(data["Syed Dawood"].keys())
ws.append(headings)

for std in data:
    marks = data[std].values()
    ws.append([std] + list(marks))

char = get_column_letter(len(data["Syed Dawood"]) + 2)
for col in range(2, len(data)):
    char2 = get_column_letter(col)
    for row in range(2, len(data["Syed Dawood"]) + 2):
        ws[char2 + str(len(data["Syed Dawood"]) + 2)] = f'=SUM({char}{row}:{char}{len(data["Syed Dawood"])+1})'

wb.save('MarksOfStudents.xlsx')