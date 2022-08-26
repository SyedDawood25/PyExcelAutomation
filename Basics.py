from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

'''The Basics'''
# wb = load_workbook('wb1.xlsx')
# ws = wb.active

# ws['A4'] = 4
# ws['B4'] = "Hamza"
# ws['C4'] = 20

# wb.save('wb1.xlsx')

'''Beginner (Creating Sheets)'''
# wb = load_workbook('wb1.xlsx')
# wb.create_sheet('Test')
# ws = wb['Test']

# print(ws)
# wb.save('wb1.xlsx')

'''Intermediate (Creating Workbook)'''
# wb = Workbook()
# ws = wb.active
# ws.title = 'Test'

# ws.append([1, 'Syed Dawood', 19])
# ws.append([2, 'Moiz Habib', 19])
# ws.append([3, 'Saad Rasheed', 20])
# ws.append([4, 'Ali Muhammad', 21])

# wb.save('wb2.xlsx')
'''(Accessing Multiple Cells)'''
# wb = load_workbook('wb2.xlsx')
# ws = wb.active
# for row in range(1, 5):
#     for col in range(1, 4):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)

'''Advanced (Formatting Elements)'''
# Merging/Unmerging Cells
# wb = load_workbook('wb1.xlsx')
# ws = wb.active

# ws.merge_cells('A1:D4')
# ws.unmerge_cells('A1:D4')

# wb.save('wb1.xlsx')

# Inserting Rows and Cols
# wb = load_workbook('wb1.xlsx')
# ws = wb.active

# ws.insert_rows(1)
# ws.insert_cols(1)

# wb.save('wb1.xlsx')
'''(Moving/Shifting rows or columns)'''
# wb = load_workbook('wb1.xlsx')
# ws = wb.active

# ws.move_range('F6', rows=-1)

# wb.save('wb1.xlsx')